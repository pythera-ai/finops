import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import io
import boto3
from typing import Dict, Any, Optional
import tempfile


def main(
    aws_access_key_id: str,
    aws_secret_access_key: str,
    input_s3_bucket: str,
    input_s3_key: str,
    output_s3_bucket: str,
    s3_region: str = None,
) -> Dict[str, Any]:
    """
    Convert CSV file from S3 to Excel XLSX with VPBank Securities formatting and upload back to S3

    Args:
        s3_resource: AWS S3 credentials and configuration
        input_s3_bucket: S3 bucket containing the input CSV file
        input_s3_key: S3 key (path) to the input CSV file
        output_s3_bucket: S3 bucket for the output Excel file
        output_s3_key: S3 key (path) for the output Excel file (optional, auto-generated if not provided)

    Returns:
        Dict containing conversion results and S3 paths
    """

    try:
        # Tạo S3 client
        s3_client = boto3.client(
            "s3",
            aws_access_key_id="aws_access_key_id",
            aws_secret_access_key="aws_secret_access_key",
            region_name=s3_region,
        )

        # Lấy tên file từ input key và đổi extension
        base_name = os.path.splitext(os.path.basename(input_s3_key))[0]
        output_s3_key = f"{base_name}.xlsx"

        # Tạo temporary file để download CSV từ S3
        with tempfile.NamedTemporaryFile(
            mode="w+", suffix=".csv", delete=False, encoding="utf-8"
        ) as temp_csv:
            temp_csv_path = temp_csv.name

        # Download CSV file từ S3
        print(f"📥 Downloading CSV from s3://{input_s3_bucket}/{input_s3_key}")
        s3_client.download_file(input_s3_bucket, input_s3_key, temp_csv_path)

        # Đọc và xử lý file CSV
        with open(temp_csv_path, "r", encoding="utf-8") as file:
            lines = file.readlines()

        # Xử lý dữ liệu
        data_sections = []
        current_section = []
        title_row = None

        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                if current_section:
                    data_sections.append(current_section)
                    current_section = []
                continue

            # Dòng đầu tiên là tiêu đề
            if i == 0:
                title_row = line
                continue

            current_section.append(line)

        # Thêm section cuối nếu có
        if current_section:
            data_sections.append(current_section)

        # Tạo workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Summary"

        # Định nghĩa màu sắc theo VPBank Securities
        colors = {
            "title_bg": "4CAF50",  # Xanh VPBank cho tiêu đề
            "title_text": "FFFFFF",  # Trắng cho text tiêu đề
            "header_bg": "43A047",  # Xanh đậm VPBank cho header
            "header_text": "FFFFFF",  # Trắng cho text header
            "increase": "4CAF50",  # Xanh lá VPBank cho tăng
            "decrease": "FF0000",  # Đỏ cho giảm
            "neutral": "000000",  # Đen cho trung tính
            "special_bg": "E8F5E8",  # Xanh nhạt VPBank cho background
            "vpbank_green": "43A047",  # Xanh đậm VPBank
        }

        current_row = 1

        # Thêm tiêu đề
        if title_row:
            ws.merge_cells(f"A{current_row}:G{current_row}")
            cell = ws[f"A{current_row}"]
            cell.value = title_row
            cell.font = Font(
                name="Times New Roman", bold=True, size=14, color=colors["title_text"]
            )
            cell.fill = PatternFill(
                start_color=colors["title_bg"],
                end_color=colors["title_bg"],
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[current_row].height = 30
            current_row += 1

        # Xử lý từng section
        for section in data_sections:
            if not section:
                continue

            first_line = section[0]

            # Kiểm tra nếu là section chỉ số chứng khoán (có header)
            if "CHỈ SỐ" in first_line and "ĐIỂM" in first_line:
                # Tạo header
                headers = [
                    "CHỈ SỐ",
                    "ĐIỂM",
                    "(+/-)",
                    "(+/- %)",
                    "KLGD( triệu cp)",
                    "GTGD ( tỷ)",
                    "CP tăng/giảm",
                ]

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = Font(
                        name="Times New Roman",
                        bold=True,
                        size=11,
                        color=colors["header_text"],
                    )
                    cell.fill = PatternFill(
                        start_color=colors["header_bg"],
                        end_color=colors["header_bg"],
                        fill_type="solid",
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                ws.row_dimensions[current_row].height = 25
                current_row += 1

                # Xử lý dữ liệu chỉ số
                for line in section[1:]:
                    if not line.strip():
                        continue

                    # Parse dòng dữ liệu chỉ số
                    parts = line.split(",")
                    if len(parts) >= 6:
                        # Tên chỉ số
                        cell = ws.cell(row=current_row, column=1, value=parts[0])
                        cell.font = Font(
                            name="Times New Roman",
                            bold=True,
                            size=11,
                            color=colors["decrease"],
                        )
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                        # Điểm
                        cell = ws.cell(row=current_row, column=2, value=parts[1])
                        cell.font = Font(
                            name="Times New Roman", size=11, color=colors["decrease"]
                        )
                        cell.alignment = Alignment(
                            horizontal="right", vertical="center"
                        )

                        # (+/-)
                        change_val = parts[2]
                        cell = ws.cell(row=current_row, column=3, value=change_val)
                        if "-" in change_val:
                            cell.font = Font(
                                name="Times New Roman",
                                size=11,
                                color=colors["decrease"],
                            )
                        else:
                            cell.font = Font(
                                name="Times New Roman",
                                size=11,
                                color=colors["increase"],
                            )
                        cell.alignment = Alignment(
                            horizontal="right", vertical="center"
                        )

                        # (+/- %)
                        percent_val = parts[3]
                        cell = ws.cell(row=current_row, column=4, value=percent_val)
                        if "-" in percent_val:
                            cell.font = Font(
                                name="Times New Roman",
                                size=11,
                                color=colors["decrease"],
                            )
                        else:
                            cell.font = Font(
                                name="Times New Roman",
                                size=11,
                                color=colors["increase"],
                            )
                        cell.alignment = Alignment(
                            horizontal="right", vertical="center"
                        )

                        # KLGD
                        cell = ws.cell(row=current_row, column=5, value=parts[4])
                        cell.font = Font(name="Times New Roman", size=11)
                        cell.alignment = Alignment(
                            horizontal="right", vertical="center"
                        )

                        # GTGD
                        cell = ws.cell(row=current_row, column=6, value=parts[5])
                        cell.font = Font(name="Times New Roman", size=11)
                        cell.alignment = Alignment(
                            horizontal="right", vertical="center"
                        )

                        # CP tăng/giảm
                        if len(parts) > 6:
                            cp_change = parts[6]
                            cell = ws.cell(row=current_row, column=7, value=cp_change)

                            # Tách và tô màu từng phần
                            if "|" in cp_change:
                                # Format: 69|33|263 -> 69 xanh, 33 đen, 263 đỏ
                                parts_cp = cp_change.split("|")
                                formatted_text = (
                                    f"{parts_cp[0]}|{parts_cp[1]}|{parts_cp[2]}"
                                )
                                cell.value = formatted_text
                                cell.font = Font(name="Times New Roman", size=11)
                            else:
                                cell.font = Font(name="Times New Roman", size=11)
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )

                        ws.row_dimensions[current_row].height = 25
                        current_row += 1

            else:
                # Các section khác (Khối ngoại, Top mua ròng, etc.)
                for line in section:
                    if not line.strip():
                        continue

                    # Làm sạch dữ liệu: bỏ dấu phẩy và ngoặc kép
                    cleaned_line = line.replace('"', "").replace(",", " ")

                    # Merge toàn bộ dòng
                    ws.merge_cells(f"A{current_row}:G{current_row}")
                    cell = ws[f"A{current_row}"]
                    cell.value = cleaned_line

                    # Xác định màu dựa trên nội dung với màu VPBank
                    if "Khối ngoại" in line:
                        if "+" in line:
                            cell.font = Font(
                                name="Times New Roman",
                                bold=True,
                                size=11,
                                color=colors["increase"],
                            )
                        else:
                            cell.font = Font(
                                name="Times New Roman",
                                bold=True,
                                size=11,
                                color=colors["decrease"],
                            )
                        cell.fill = PatternFill(
                            start_color=colors["special_bg"],
                            end_color=colors["special_bg"],
                            fill_type="solid",
                        )
                    elif "Top mua ròng" in line:
                        cell.font = Font(
                            name="Times New Roman",
                            bold=True,
                            size=11,
                            color=colors["increase"],
                        )
                        cell.fill = PatternFill(
                            start_color=colors["special_bg"],
                            end_color=colors["special_bg"],
                            fill_type="solid",
                        )
                    elif "Top bán ròng" in line:
                        cell.font = Font(
                            name="Times New Roman",
                            bold=True,
                            size=11,
                            color=colors["decrease"],
                        )
                    elif "Khối tự doanh" in line:
                        if "+" in line:
                            cell.font = Font(
                                name="Times New Roman",
                                bold=True,
                                size=11,
                                color=colors["increase"],
                            )
                        else:
                            cell.font = Font(
                                name="Times New Roman",
                                bold=True,
                                size=11,
                                color=colors["decrease"],
                            )
                        cell.fill = PatternFill(
                            start_color=colors["special_bg"],
                            end_color=colors["special_bg"],
                            fill_type="solid",
                        )
                    elif "Nhóm ngành" in line:
                        cell.font = Font(
                            name="Times New Roman", size=11, color=colors["neutral"]
                        )
                        cell.fill = PatternFill(
                            start_color=colors["special_bg"],
                            end_color=colors["special_bg"],
                            fill_type="solid",
                        )
                    elif "Cổ phiếu" in line:
                        cell.font = Font(
                            name="Times New Roman", size=11, color=colors["neutral"]
                        )
                    elif "Tác động tăng" in line:
                        cell.font = Font(
                            name="Times New Roman", size=11, color=colors["increase"]
                        )
                    elif "Tác động giảm" in line:
                        cell.font = Font(
                            name="Times New Roman", size=11, color=colors["decrease"]
                        )
                    else:
                        cell.font = Font(
                            name="Times New Roman", size=11, color=colors["neutral"]
                        )

                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    ws.row_dimensions[current_row].height = 25
                    current_row += 1

        # Thêm border cho toàn bộ bảng
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(
            min_row=1, max_row=current_row - 1, min_col=1, max_col=7
        ):
            for cell in row:
                cell.border = thin_border

        # Điều chỉnh độ rộng cột
        column_widths = [15, 12, 10, 10, 18, 15, 20]  # Độ rộng tùy chỉnh cho từng cột
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Tạo temporary file để lưu Excel
        with tempfile.NamedTemporaryFile(
            mode="w+b", suffix=".xlsx", delete=False
        ) as temp_excel:
            temp_excel_path = temp_excel.name

        # Lưu Excel file
        wb.save(temp_excel_path)

        # Upload Excel file lên S3
        print(f"📤 Uploading Excel to s3://{output_s3_bucket}/{output_s3_key}")
        s3_client.upload_file(temp_excel_path, output_s3_bucket, output_s3_key)

        # Lấy file size
        file_size = os.path.getsize(temp_excel_path)

        # Cleanup temporary files
        os.unlink(temp_csv_path)
        os.unlink(temp_excel_path)

        return {
            "success": True,
            "message": "✅ CSV đã được convert thành công sang Excel và upload lên S3!",
            "input_s3_path": f"s3://{input_s3_bucket}/{input_s3_key}",
            "output_s3_path": f"s3://{output_s3_bucket}/{output_s3_key}",
            "output_s3_bucket": output_s3_bucket,
            "output_s3_key": output_s3_key,
            "rows_processed": current_row - 1,
            "file_size_bytes": file_size,
            "formatting": {
                "title_background": colors["title_bg"],
                "header_background": colors["header_bg"],
                "font": "Times New Roman",
                "vpbank_style": True,
            },
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"❌ Lỗi khi convert file: {str(e)}",
            "message": "Vui lòng kiểm tra lại đường dẫn S3 và quyền truy cập",
            "input_s3_path": f"s3://{input_s3_bucket}/{input_s3_key}",
            "output_s3_path": None,
        }


def validate_s3_paths(
    s3_resource: dict, input_bucket: str, input_key: str, output_bucket: str
) -> Dict[str, Any]:
    """
    Validate S3 paths and permissions before processing
    """
    try:
        s3_client = boto3.client(
            "s3",
            aws_access_key_id="aws_access_key_id",
            aws_secret_access_key="aws_secret_access_key",
            region_name=s3_resource.get("region", "us-east-1"),
        )

        # Check if input file exists
        try:
            s3_client.head_object(Bucket=input_bucket, Key=input_key)
            input_exists = True
        except:
            input_exists = False

        # Check if output bucket exists
        try:
            s3_client.head_bucket(Bucket=output_bucket)
            output_bucket_exists = True
        except:
            output_bucket_exists = False

        return {
            "input_file_exists": input_exists,
            "output_bucket_exists": output_bucket_exists,
            "validation_passed": input_exists and output_bucket_exists,
            "input_s3_path": f"s3://{input_bucket}/{input_key}",
            "output_s3_bucket": f"s3://{output_bucket}",
        }

    except Exception as e:
        return {"validation_passed": False, "error": str(e)}
