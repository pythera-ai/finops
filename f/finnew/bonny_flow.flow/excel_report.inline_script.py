import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import os
import boto3
from typing import Dict, Any
import tempfile
from datetime import datetime


def create_enhanced_excel_report(data: Dict[str, Any]) -> str:
    """Create Excel report matching customer-accepted visual format"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Summary"

    # Colors
    colors = {
        "title_bg": "4CAF50",
        "title_text": "FF0000",
        "header_bg": "FAC090",
        "header_text": "000000",
        "increase": "008000",
        "decrease": "FF0000",
        "neutral": "000000",
        "yellow": "FFD700",
        "light_gray": "F0F0F0",
        "section_bg": "E8E8E8",
    }

    # Fonts
    fonts = {
        "title": Font(
            name="Times New Roman", bold=True, size=14, color=colors["title_text"]
        ),
        "header": Font(
            name="Times New Roman", bold=True, size=11, color=colors["header_text"]
        ),
        "body": Font(name="Times New Roman", size=11, color=colors["neutral"]),
        "body_bold": Font(
            name="Times New Roman", bold=True, size=11, color=colors["neutral"]
        ),
        "increase": Font(
            name="Times New Roman", bold=True, size=11, color=colors["increase"]
        ),
        "decrease": Font(
            name="Times New Roman", bold=True, size=11, color=colors["decrease"]
        ),
    }

    # Border
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def apply_border_to_merged_range(
        worksheet, start_row, start_col, end_row, end_col, border_style
    ):
        """Apply borders to all cells in a merged range"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                worksheet.cell(row=row, column=col).border = border_style

    def get_color_based_on_percentage(value_str: str) -> str:
        """Determine color based on percentage value"""
        try:
            clean_value = str(value_str).replace("%", "").replace(",", ".")
            return (
                "decrease"
                if clean_value.startswith("-") or float(clean_value) < 0
                else "increase"
            )
        except:
            return "neutral"

    def get_font_for_color_rule(color_type: str) -> Font:
        """Get appropriate font based on color rule"""
        return fonts.get(color_type, fonts["body_bold"])

    def format_number_with_sign(value: float) -> str:
        """Format number with appropriate arrow icon"""
        if value > 0:
            return f"↑{value:.2f}"
        elif value < 0:
            return f"↓{abs(value):.2f}"
        else:
            return f"↔{value:.2f}"

    current_row = 1

    # Title row
    report_date = datetime.now().strftime("%d/%m/%Y")
    title = f"TỔNG KẾT GIAO DỊCH PHIÊN {report_date}"

    ws.merge_cells(f"A{current_row}:I{current_row}")
    title_cell = ws[f"A{current_row}"]
    title_cell.value = title
    title_cell.font = fonts["title"]
    title_cell.fill = PatternFill(
        start_color=colors["title_bg"], end_color=colors["title_bg"], fill_type="solid"
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border
    apply_border_to_merged_range(ws, current_row, 1, current_row, 9, border)
    ws.row_dimensions[current_row].height = 35
    current_row += 1

    # Headers
    headers = [
        "CHỈ SỐ",
        "ĐIỂM",
        "(+/-)",
        "(+/- %)",
        "KLGD( triệu cp)",
        "(+/-%KLGD)",
        "GTGD ( tỷ)",
        "(+/-%GTGD)",
        "CP tăng/giảm",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = fonts["header"]
        cell.fill = PatternFill(
            start_color=colors["header_bg"],
            end_color=colors["header_bg"],
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[current_row].height = 25
    current_row += 1

    # Index data
    index_data = data.get("index_summary", [])
    index_mapping_ordered = [
        ("VNINDEX", "VNINDEX"),
        ("VN30", "VN30"),
        ("HNXIndex", "HNXINDEX"),
        ("HNX30", "HNX30"),
        ("HNXUpcomIndex", "UPCOM"),
    ]
    index_dict = {item.get("indexId", "Unknown"): item for item in index_data}

    row_count = 0
    for data_label, display_name in index_mapping_ordered:
        item = index_dict.get(data_label)
        if not item:
            continue

        bg_color = colors["light_gray"] if row_count % 2 == 0 else "FFFFFF"
        change_percent = item.get("changePercent", 0)
        color_rule = get_color_based_on_percentage(str(change_percent))

        # Data columns
        columns_data = [
            (display_name, "left", get_font_for_color_rule(color_rule)),
            (
                f"{float(item.get('indexValue', 0)):.2f}",
                "right",
                get_font_for_color_rule(color_rule),
            ),
            (
                format_number_with_sign(float(item.get("change", 0))),
                "right",
                get_font_for_color_rule(
                    get_color_based_on_percentage(str(item.get("change", 0)))
                ),
            ),
            (
                format_number_with_sign(float(change_percent)),
                "right",
                get_font_for_color_rule(
                    get_color_based_on_percentage(str(change_percent))
                ),
            ),
            (f"{float(item.get('allQty', 0)):.2f}", "right", fonts["body"]),
            (
                format_number_with_sign(float(item.get("klgd_change_percent", 0))),
                "right",
                get_font_for_color_rule(
                    get_color_based_on_percentage(
                        str(item.get("klgd_change_percent", 0))
                    )
                ),
            ),
            (f"{float(item.get('allValue', 0)):.2f}", "right", fonts["body"]),
            (
                format_number_with_sign(float(item.get("gtdg_change_percent", 0))),
                "right",
                get_font_for_color_rule(
                    get_color_based_on_percentage(
                        str(item.get("gtdg_change_percent", 0))
                    )
                ),
            ),
        ]

        for col_idx, (value, alignment, font) in enumerate(columns_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = font
            cell.fill = PatternFill(
                start_color=bg_color, end_color=bg_color, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal=alignment, vertical="center")
            cell.border = border

        # CP tăng/giảm column with rich text
        advances = item.get("advances", 0)
        nochanges = item.get("nochanges", 0)
        declines = item.get("declines", 0)
        rich_text_parts = []

        if advances > 0:
            rich_text_parts.append(
                TextBlock(
                    InlineFont(rFont="Times New Roman", sz=11, color="008000", b=True),
                    f"↑{advances}",
                )
            )
        if len(rich_text_parts) > 0 and (nochanges > 0 or declines > 0):
            rich_text_parts.append(
                TextBlock(
                    InlineFont(rFont="Times New Roman", sz=11, color="000000"), "|"
                )
            )
        if nochanges > 0:
            rich_text_parts.append(
                TextBlock(
                    InlineFont(rFont="Times New Roman", sz=11, color="FFD700", b=True),
                    f"↔{nochanges}",
                )
            )
        if len(rich_text_parts) > 0 and declines > 0:
            rich_text_parts.append(
                TextBlock(
                    InlineFont(rFont="Times New Roman", sz=11, color="000000"), "|"
                )
            )
        if declines > 0:
            rich_text_parts.append(
                TextBlock(
                    InlineFont(rFont="Times New Roman", sz=11, color="FF0000", b=True),
                    f"↓{declines}",
                )
            )

        cell = ws.cell(row=current_row, column=9)
        cell.value = CellRichText(rich_text_parts) if rich_text_parts else "N/A"
        cell.fill = PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

        ws.row_dimensions[current_row].height = 20
        current_row += 1
        row_count += 1

    # Empty row
    ws.merge_cells(f"A{current_row}:I{current_row}")
    ws[f"A{current_row}"].border = border
    apply_border_to_merged_range(ws, current_row, 1, current_row, 9, border)
    current_row += 1

    # Information sections
    sections = []

    # Khối ngoại
    kn_data = data.get("khoi_ngoai", {})
    if kn_data:
        vol = kn_data.get("vol", 0)
        net_value = kn_data.get("net_value", 0)
        vol_text = f"↑{vol:.2f}" if vol >= 0 else f"↓{abs(vol):.2f}"
        net_text = (
            f"↑{net_value:.2f} tỷ đồng"
            if net_value >= 0
            else f"↓{abs(net_value):.2f} tỷ đồng"
        )
        sections.append(
            (
                f"Khối ngoại: {vol_text} triệu cp {net_text}",
                get_font_for_color_rule(get_color_based_on_percentage(str(net_value))),
            )
        )

    # Top mua/bán ròng with rich text
    top_net = data.get("top_netforeign", {})
    if top_net:
        buy_stocks = top_net.get("buy", [])
        sell_stocks = top_net.get("sell", [])

        # Top mua ròng
        buy_rich_text = [
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=11, color="008000", b=True),
                "Top mua ròng: ",
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=11, color="008000", b=False),
                ", ".join(buy_stocks) if buy_stocks else "N/A",
            ),
        ]
        sections.append((CellRichText(buy_rich_text), None))

        # Top bán ròng
        sell_rich_text = [
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=11, color="FF0000", b=True),
                "Top bán ròng: ",
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=11, color="FF0000", b=False),
                ", ".join(sell_stocks) if sell_stocks else "N/A",
            ),
        ]
        sections.append((CellRichText(sell_rich_text), None))

    # Khối tự doanh
    ktd_value = data.get("khoi_tu_doanh", 0)
    ktd_text = (
        f"↑{ktd_value:.0f} tỷ đồng"
        if ktd_value >= 0
        else f"↓{abs(ktd_value):.0f} tỷ đồng"
    )
    sections.append(
        (
            f"Khối tự doanh: {ktd_text}",
            get_font_for_color_rule(get_color_based_on_percentage(str(ktd_value))),
        )
    )

    # Nhóm ngành nổi bật and Cổ phiếu tâm điểm with rich text
    for section_name, section_data_key in [
        ("Nhóm ngành nổi bật", "top_sectors"),
        ("Cổ phiếu tâm điểm", "top_interested"),
    ]:
        section_data = data.get(section_data_key, [])
        content_text = ", ".join(section_data) if section_data else "N/A"
        rich_text = [
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=11, color="000000", b=True),
                f"{section_name}: ",
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=11, color="000000", b=False),
                content_text,
            ),
        ]
        sections.append((CellRichText(rich_text), None))

    # Impact sections with rich text
    for impact_key, color, sign in [
        ("impact_up", "008000", "+"),
        ("impact_down", "FF0000", "-"),
    ]:
        impact_data = data.get(impact_key, {})
        if impact_data:
            total = impact_data.get("total", 0)
            stocks = impact_data.get("stock_code", [])
            content_text = ", ".join(stocks) if stocks else "N/A"
            section_name = f"Tác động {'tăng' if sign == '+' else 'giảm'} ({sign}{abs(total):.2f}): "
            rich_text = [
                TextBlock(
                    InlineFont(rFont="Times New Roman", sz=11, color=color, b=True),
                    section_name,
                ),
                TextBlock(
                    InlineFont(rFont="Times New Roman", sz=11, color=color, b=False),
                    content_text,
                ),
            ]
            sections.append((CellRichText(rich_text), None))

    # Apply sections
    for section_content, section_font in sections:
        ws.merge_cells(f"A{current_row}:I{current_row}")
        cell = ws[f"A{current_row}"]
        cell.value = section_content
        if section_font:
            cell.font = section_font
        cell.fill = PatternFill(
            start_color=colors["section_bg"],
            end_color=colors["section_bg"],
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        apply_border_to_merged_range(ws, current_row, 1, current_row, 9, border)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Column widths
    column_widths = [12, 10, 8, 8, 15, 12, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return wb


def main(
    data: Dict[str, Any],
    s3_bucket_name: str,
    s3_excel_key: str,
    s3_region: str = "ap-southeast-2",
    aws_access_key_id: str = None,
    aws_secret_access_key: str = None,
) -> Dict[str, Any]:
    """Generate Excel report and upload to S3"""

    try:
        s3_client = boto3.client(
            "s3",
            region_name=s3_region,
            aws_access_key_id=aws_access_key_id,
            aws_secret_access_key=aws_secret_access_key,
        )

        wb = create_enhanced_excel_report(data)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            wb.save(temp_file.name)
            temp_file_path = temp_file.name

        s3_client.upload_file(temp_file_path, s3_bucket_name, s3_excel_key)
        file_size = os.path.getsize(temp_file_path)
        os.unlink(temp_file_path)

        return {
            "success": True,
            "message": "Excel report generated successfully",
            "s3_location": f"s3://{s3_bucket_name}/{s3_excel_key}",
            "file_size_bytes": file_size,
            "timestamp": datetime.now().isoformat(),
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Excel generation failed: {str(e)}",
            "error": str(e),
            "timestamp": datetime.now().isoformat(),
        }
